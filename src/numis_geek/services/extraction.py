"""Spec 38 — orchestrates LLM extraction jobs.

Three public entrypoints:

- `create_and_run` — builds a job and runs the LLM call inline (sync mode).
- `confirm_extraction` — applies the extracted (or user-edited) payload to
  the domain model and marks the source pendency resolved.
- `reject_extraction` — records a manual rejection.

V1 ships **sync** extraction (no background worker). Files are small
enough that a 10-30s POST is fine. Move to async + APScheduler when
batch uploads become a thing.
"""
from __future__ import annotations

import json
import re
import uuid
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

from pydantic import ValidationError
from sqlalchemy import func
from sqlalchemy.orm import Session

from numis_geek.models.account import Account, Currency
from numis_geek.models.asset import Asset, AssetClass
from numis_geek.models.asset_movement import AssetMovement, AssetMovementType
from numis_geek.models.attachment import Attachment, AttachmentKind
from numis_geek.models.audit_log import AuditLog  # noqa: F401  (used implicitly via AuditService)
from numis_geek.models.distribution import Distribution, DistributionType
from numis_geek.models.external import ExternalSource
from numis_geek.models.extraction_job import (
    ExtractionJob, ExtractionSourceHint, ExtractionStatus,
)
from numis_geek.models.financial_institution import FinancialInstitution
from numis_geek.models.portfolio_snapshot import (
    PortfolioSnapshot, PortfolioSnapshotItem, SnapshotPendency, SnapshotStatus,
)
from numis_geek.models.user import User
from numis_geek.services import attachment_storage
from numis_geek.services.audit import AuditService
from numis_geek.services.extraction_templates import Template, template_for
from numis_geek.services.fx import resolve_fx_rate
from numis_geek.integrations.llm import (
    DEFAULT_MODEL, LLMClient, get_llm_client, parse_json_block,
)


# ── data classes ─────────────────────────────────────────────────────────────

@dataclass
class BulkApplyDetail:
    """Spec 48 — categorized result of bulk extract apply.

    Each entry is a plain dict so the route layer can serialize via Pydantic
    without re-mapping field names.

    For PREVIEW (read-only): `applied` becomes the prospective list — same
    shape, but `new_price`/`previous_price` reflect what WOULD happen.
    """
    applied: list[dict]              # matched + pendency (resolved for apply, prospective for preview)
    matched_no_pendency: list[dict]  # asset exists but no open pendency in snapshot
    orphan: list[dict]               # extract line had no matching asset
    pendency_not_in_extract: list[dict]  # snapshot pendency the extract didn't cover
    # Spec 57 follow-up — asset matched but price_source is automated
    # (BRAPI/FINNHUB/...); extract is ignored. UI shows this so the user
    # knows the line was recognized but no action is needed.
    auto_skipped: list[dict] = None  # populated only by preview/apply paths that classify

    def __post_init__(self):
        if self.auto_skipped is None:
            self.auto_skipped = []


@dataclass
class ExtractionApplyResult:
    applied_count: int
    skipped_count: int
    errors: list[str]
    bulk_detail: BulkApplyDetail | None = None


class ExtractionError(Exception):
    pass


# ── disambiguator ────────────────────────────────────────────────────────────

# Datas no notes vêm em três formatos vistos em extratos: DD/MM/YYYY,
# YYYY-MM-DD, MM/YYYY. Captura o mais específico que encontrar.
_DATE_PATTERNS = (
    re.compile(r"\b(\d{4})-(\d{2})-(\d{2})\b"),       # 2034-08-16
    re.compile(r"\b(\d{2})/(\d{2})/(\d{4})\b"),       # 16/08/2034
    re.compile(r"\b(\d{2})/(\d{4})\b"),               # 08/2034
)


def _maturity_from_notes(notes: str | None) -> str | None:
    """Return an ISO-ish maturity tag (`2034-08-16` or `2034-08`) extracted
    from the free-text notes, or None if no date pattern matches."""
    if not notes:
        return None
    text = notes
    if m := _DATE_PATTERNS[0].search(text):
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    if m := _DATE_PATTERNS[1].search(text):
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    if m := _DATE_PATTERNS[2].search(text):
        return f"{m.group(2)}-{m.group(1)}"
    return None


def _disambiguate_duplicate_tickers(validated: Any) -> None:
    """Spec 57 defensive layer: when two positions returned by the LLM share
    a ticker_raw (e.g. 10x 'United States of America' for distinct
    Treasury maturities), append the maturity extracted from `notes` to
    make each ticker_raw unique. Mutates `validated.positions` in place.

    Safe no-op when:
    - the output model has no `positions` attribute (other source_hints)
    - all ticker_raws are already unique
    - duplicate has no `notes` to mine a date from (left as-is; downstream
      still flags it as orphan, but at least the user sees ONE per row)
    """
    positions = getattr(validated, "positions", None)
    if not positions:
        return
    counts = Counter(p.ticker_raw for p in positions if p.ticker_raw)
    dups = {t for t, n in counts.items() if n > 1}
    if not dups:
        return
    seen: set[str] = set(t for t in counts if t not in dups)
    for pos in positions:
        if pos.ticker_raw not in dups:
            continue
        maturity = _maturity_from_notes(getattr(pos, "notes", None))
        if not maturity:
            continue
        new_raw = f"{pos.ticker_raw} {maturity}"
        # Bump suffix if a date collision still occurs (rare).
        suffix = 2
        candidate = new_raw
        while candidate in seen:
            candidate = f"{new_raw} #{suffix}"
            suffix += 1
        pos.ticker_raw = candidate
        seen.add(candidate)


# ── create + run (sync) ──────────────────────────────────────────────────────

def create_and_run(
    db: Session,
    *,
    workspace_id: str,
    attachment_id: str,
    source_hint: ExtractionSourceHint,
    pendency_id: str | None = None,
    snapshot_id: str | None = None,
    asset_id: str | None = None,
    institution_id: str | None = None,
    user_id: str | None = None,
    user_email: str | None = None,
) -> ExtractionJob:
    """Create an ExtractionJob and run the LLM inline (Spec 38 V1 sync mode).

    `institution_id` (Spec 58) scopes the bulk-extract apply pool to a
    single FI when set — both display name AND candidate Asset filter.
    """
    attachment = db.get(Attachment, attachment_id)
    if attachment is None or not attachment.is_active:
        raise ExtractionError(f"Attachment {attachment_id} not found")
    if attachment.workspace_id != workspace_id:
        raise ExtractionError("Attachment workspace mismatch")

    job = ExtractionJob(
        id=str(uuid.uuid4()),
        workspace_id=workspace_id,
        snapshot_id=snapshot_id,
        pendency_id=pendency_id,
        asset_id=asset_id,
        institution_id=institution_id,
        attachment_id=attachment_id,
        source_hint=source_hint,
        status=ExtractionStatus.PENDING,
        created_at=datetime.now(timezone.utc),
    )
    db.add(job)
    db.flush()

    if user_id or user_email:
        AuditService(db).log(
            user_email=user_email or "system",
            action="extraction.created",
            workspace_id=workspace_id,
            user_id=user_id,
            resource_type="extraction_job",
            resource_id=job.id,
            details={
                "attachment_id": attachment_id,
                "source_hint": source_hint.value,
                "pendency_id": pendency_id,
            },
        )

    run_extraction(db, job_id=job.id)
    return db.get(ExtractionJob, job.id)


def run_extraction(db: Session, *, job_id: str) -> ExtractionJob:
    """Read the attachment, call the LLM, validate the JSON, persist."""
    job = db.get(ExtractionJob, job_id)
    if job is None:
        raise ExtractionError(f"ExtractionJob {job_id} not found")
    if job.status not in (ExtractionStatus.PENDING, ExtractionStatus.FAILED):
        raise ExtractionError(
            f"Cannot run job in status {job.status.value}",
        )

    attachment = db.get(Attachment, job.attachment_id)
    if attachment is None:
        job.status = ExtractionStatus.FAILED
        job.error_message = "Attachment row vanished"
        db.flush()
        return job

    # Spec 58 Stage 3 — pick per-FI template when the job is scoped to a
    # specific institution. Avenue → BROKER_POSITION_AVENUE etc.
    fi_short_name: str | None = None
    if job.institution_id:
        fi = db.get(FinancialInstitution, job.institution_id)
        fi_short_name = fi.short_name if fi else None

    # Spec 58 Stage 4 — if a deterministic Python parser exists for this
    # (fi, hint, mime) combination, skip the LLM entirely. 100% accurate,
    # zero token cost. Falls back to LLM when no parser matches.
    from numis_geek.services.extraction_parsers import parser_for as _parser_for
    det_parser = _parser_for(
        institution_short_name=fi_short_name,
        source_hint=job.source_hint,
        mime_type=attachment.mime_type,
    )

    job.status = ExtractionStatus.RUNNING
    job.started_at = datetime.now(timezone.utc)
    db.flush()

    if det_parser is not None:
        try:
            blob = Path(attachment_storage.absolute_path_for(attachment)).read_bytes()
            payload_dict = det_parser(blob)
            job.extracted_json = payload_dict
            job.prompt_version = f"deterministic:{det_parser.__name__}"
            job.model = None
            job.input_tokens = 0
            job.output_tokens = 0
            job.cost_usd = Decimal("0")
            job.confidence = Decimal("1.00")
            job.status = ExtractionStatus.EXTRACTED
            job.completed_at = datetime.now(timezone.utc)
            job.error_message = None
        except Exception as exc:
            job.status = ExtractionStatus.FAILED
            job.error_message = f"Deterministic parser failed: {exc}"
            job.completed_at = datetime.now(timezone.utc)
        db.flush()
        return job

    template: Template = template_for(
        job.source_hint, institution_short_name=fi_short_name,
    )
    job.prompt_version = template.version
    db.flush()

    # Para BROKER_INCOME scoped a snapshot, passa o período pra o prompt
    # (o filtro autoritativo é server-side em _classify_bulk_income).
    period_label: str | None = None
    if (
        job.source_hint == ExtractionSourceHint.BROKER_INCOME
        and job.snapshot_id
    ):
        snap = db.get(PortfolioSnapshot, job.snapshot_id)
        if snap:
            period_start, period_end = _income_period_bounds(snap)
            period_label = (
                f"{period_start.isoformat()} a {period_end.isoformat()}"
            )

    try:
        client: LLMClient = get_llm_client(db)
        payload = _read_attachment_payload(attachment)
        call = client.call(
            system=template.system,
            user_text=_user_message(template, payload, period_label=period_label),
            image_bytes=payload.image_bytes,
            image_mime=payload.image_mime,
            image_parts=payload.image_parts,
            model=DEFAULT_MODEL,
        )
        parsed_obj = parse_json_block(call.text)
        validated = template.output_model.model_validate(parsed_obj)
        _disambiguate_duplicate_tickers(validated)

        job.extracted_json = validated.model_dump()
        job.model = call.model
        job.input_tokens = call.input_tokens
        job.output_tokens = call.output_tokens
        job.cost_usd = call.cost_usd()
        job.confidence = _overall_confidence(parsed_obj)
        job.status = ExtractionStatus.EXTRACTED
        job.completed_at = datetime.now(timezone.utc)
        job.error_message = None
    except (ValidationError, ValueError, json.JSONDecodeError) as exc:
        job.status = ExtractionStatus.FAILED
        job.error_message = f"Parse/validate: {exc}"
        job.completed_at = datetime.now(timezone.utc)
    except Exception as exc:  # pragma: no cover - generic backstop
        job.status = ExtractionStatus.FAILED
        job.error_message = str(exc)
        job.completed_at = datetime.now(timezone.utc)

    db.flush()
    return job


@dataclass
class _Payload:
    text: str | None
    image_bytes: bytes | None
    image_mime: str | None
    image_parts: list[tuple[bytes, str | None]] | None = None


_ANTHROPIC_IMAGE_MAX_DIM = 8000  # Anthropic's hard limit on either side.


def _split_image_for_anthropic(
    blob: bytes, mime_type: str | None,
) -> list[tuple[bytes, str | None]]:
    """Return a list of (bytes, mime) tiles that all fit Anthropic's bounds.

    For images within the 8000×8000 limit, returns `[(blob, mime)]` unchanged.
    For taller / wider stitched screenshots we slice into tiles (vertical
    strips first, then horizontal if needed). Resolution is preserved — no
    downscale — so small text remains legible to Claude. Tiles are emitted
    as JPEG quality 92 to keep payload under Anthropic's 5MB/image cap.

    Falls back to `[(blob, mime)]` when Pillow isn't installed; the API
    will reject oversize blobs in that case, surfaced as a clear error.
    """
    try:
        from io import BytesIO
        from PIL import Image
    except ImportError:
        return [(blob, mime_type)]
    try:
        img = Image.open(BytesIO(blob))
        w, h = img.size
        if w <= _ANTHROPIC_IMAGE_MAX_DIM and h <= _ANTHROPIC_IMAGE_MAX_DIM:
            return [(blob, mime_type)]
        img = img.convert("RGB") if img.mode in ("RGBA", "P", "LA") else img
        cols = (w + _ANTHROPIC_IMAGE_MAX_DIM - 1) // _ANTHROPIC_IMAGE_MAX_DIM
        rows = (h + _ANTHROPIC_IMAGE_MAX_DIM - 1) // _ANTHROPIC_IMAGE_MAX_DIM
        tiles: list[tuple[bytes, str | None]] = []
        for row in range(rows):
            for col in range(cols):
                left = col * _ANTHROPIC_IMAGE_MAX_DIM
                upper = row * _ANTHROPIC_IMAGE_MAX_DIM
                right = min(left + _ANTHROPIC_IMAGE_MAX_DIM, w)
                lower = min(upper + _ANTHROPIC_IMAGE_MAX_DIM, h)
                buf = BytesIO()
                img.crop((left, upper, right, lower)).save(buf, format="JPEG", quality=92)
                tiles.append((buf.getvalue(), "image/jpeg"))
        return tiles
    except Exception:
        return [(blob, mime_type)]


_XLSX_MIMES = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
)


def _xlsx_to_csv_text(blob: bytes) -> str | None:
    """Convert an Excel workbook to a CSV-like textual representation that
    Claude can read. Returns None when openpyxl isn't installed or the
    workbook can't be parsed — caller falls back to raw bytes."""
    try:
        from io import BytesIO
        import csv as _csv
        from openpyxl import load_workbook
    except ImportError:
        return None
    try:
        wb = load_workbook(BytesIO(blob), read_only=True, data_only=True)
    except Exception:
        return None
    parts: list[str] = []
    for ws in wb.worksheets:
        parts.append(f"# Sheet: {ws.title}")
        from io import StringIO
        out = StringIO()
        writer = _csv.writer(out)
        for row in ws.iter_rows(values_only=True):
            if all(v is None for v in row):
                continue
            writer.writerow([("" if v is None else str(v)) for v in row])
        parts.append(out.getvalue().rstrip())
    return "\n\n".join(parts)


def _read_attachment_payload(attachment: Attachment) -> _Payload:
    """Convert the on-disk attachment into LLM input (image tiles OR text)."""
    path = attachment_storage.absolute_path_for(attachment)
    blob = Path(path).read_bytes()
    if attachment.kind == AttachmentKind.IMAGE:
        parts = _split_image_for_anthropic(blob, attachment.mime_type)
        return _Payload(
            text=None,
            image_bytes=parts[0][0] if parts else None,
            image_mime=parts[0][1] if parts else None,
            image_parts=parts,
        )
    if attachment.mime_type in _XLSX_MIMES:
        text = _xlsx_to_csv_text(blob)
        if text is not None:
            return _Payload(text=text, image_bytes=None, image_mime=None)
        # Fallback: bytes (Anthropic will likely fail, but at least we tried).
        return _Payload(text=None, image_bytes=blob, image_mime=attachment.mime_type)
    if attachment.kind == AttachmentKind.CSV:
        return _Payload(text=blob.decode("utf-8", errors="replace"), image_bytes=None, image_mime=None)
    # PDF and others — try utf-8 first, else send as base64 image-fallback.
    # V1 simplification: pass the raw bytes as the prompt for non-image
    # attachments. Real PDF-to-image conversion (pypdfium / pdf2image) is a
    # future improvement.
    try:
        return _Payload(text=blob.decode("utf-8"), image_bytes=None, image_mime=None)
    except UnicodeDecodeError:
        return _Payload(text=None, image_bytes=blob, image_mime=attachment.mime_type)


def _user_message(
    template: Template,
    payload: _Payload,
    *,
    period_label: str | None = None,
) -> str:
    parts = [template.user_prefix]
    if period_label:
        parts.append(
            f"\n\nPERÍODO DO FECHAMENTO: {period_label}. Eventos fora deste "
            "intervalo serão filtrados pelo sistema — não tente forçar."
        )
    if payload.image_parts and len(payload.image_parts) > 1:
        parts.append(
            f"\n\nNOTA: a imagem original excedia o limite de 8000px, então "
            f"está dividida em {len(payload.image_parts)} fatias sequenciais "
            f"(top-to-bottom, left-to-right). Leia todas como uma única "
            f"captura contínua antes de extrair os dados.",
        )
    if payload.text:
        parts.append("\n\n----\n")
        parts.append(payload.text[:50_000])  # safety cap
    return "".join(parts)


def _overall_confidence(parsed: Any) -> Decimal | None:
    if isinstance(parsed, dict):
        if "confidence" in parsed and isinstance(parsed["confidence"], (int, float)):
            return Decimal(str(parsed["confidence"])).quantize(Decimal("0.01"))
        # For list-based hints, average the per-row confidence.
        for list_key in ("positions", "events", "trades"):
            items = parsed.get(list_key)
            if isinstance(items, list) and items:
                values = [
                    float(it.get("confidence"))
                    for it in items
                    if isinstance(it, dict) and isinstance(it.get("confidence"), (int, float))
                ]
                if values:
                    avg = sum(values) / len(values)
                    return Decimal(str(round(avg, 2)))
    return None


# ── confirm / reject ─────────────────────────────────────────────────────────

def confirm_extraction(
    db: Session,
    *,
    job_id: str,
    user_id: str | None,
    user_email: str | None = None,
    edited_payload: dict | None = None,
    institution_short_name: str | None = None,
    manual_mappings: dict[str, str] | None = None,
    manual_prices: dict[str, float] | None = None,
    manual_modes: dict[str, str] | None = None,
) -> ExtractionApplyResult:
    """Apply the (possibly edited) extracted JSON to the domain model.

    Always resolves the linked SnapshotPendency (if any). When the job is
    bulk-scoped (snapshot_id set, pendency_id None), dispatches to the bulk
    applier which resolves N pendencies. `institution_short_name`, when set,
    restricts the bulk matching to assets of that FI (Spec 48 §1.3).
    `manual_mappings` overrides auto-matching for orphan lines: keys are
    `ticker_raw` strings from the extract, values are pendency IDs the user
    chose to apply the line against (Spec 49 hotfix — review screen lets
    user manually map orphans to open pendencies).
    """
    job = db.get(ExtractionJob, job_id)
    if job is None:
        raise ExtractionError(f"ExtractionJob {job_id} not found")
    if job.status != ExtractionStatus.EXTRACTED:
        raise ExtractionError(
            f"Job must be in EXTRACTED state (currently {job.status.value})",
        )

    payload = edited_payload if edited_payload is not None else job.extracted_json
    if payload is None:
        raise ExtractionError("No extracted JSON to apply")

    result = _apply_payload(
        db, job, payload,
        user_id=user_id, user_email=user_email,
        institution_short_name=institution_short_name,
        manual_mappings=manual_mappings,
        manual_prices=manual_prices,
        manual_modes=manual_modes,
    )

    job.user_edits = edited_payload if edited_payload is not None else None
    # Spec 49 hotfix #6 — only flip to CONFIRMED when at least 1 pendency
    # was actually resolved. Otherwise leave as EXTRACTED so the user can
    # fix manual_prices / mappings and retry without re-running the LLM.
    if result.applied_count > 0:
        job.status = ExtractionStatus.CONFIRMED
        job.confirmed_at = datetime.now(timezone.utc)
        job.confirmed_by = user_id
    db.flush()

    # Resolve the linked pendency (lazy import to avoid circular module).
    if job.pendency_id:
        from numis_geek.services import snapshot as snapshot_service
        try:
            snapshot_service.resolve_pendency(
                db,
                pendency_id=job.pendency_id,
                user_id=user_id,
                user_email=user_email,
                note=f"Extraído via LLM (job {job.id})",
            )
        except ValueError:
            # Pendency might have been deleted between extraction and confirm.
            result.errors.append(f"pendency {job.pendency_id} not found")

    AuditService(db).log(
        user_email=user_email or "system",
        action="extraction.confirmed",
        workspace_id=job.workspace_id,
        user_id=user_id,
        resource_type="extraction_job",
        resource_id=job.id,
        details={
            "applied": result.applied_count,
            "skipped": result.skipped_count,
            "errors": result.errors,
            "pendency_id": job.pendency_id,
            "institution_short_name": institution_short_name,
            "manual_mappings_count": len(manual_mappings) if manual_mappings else 0,
            # Spec 49 hotfix — surface LLM usage in audit so cost can be
            # tracked by aggregating audit rows in addition to inspecting
            # individual extraction_job rows.
            "model": job.model,
            "input_tokens": job.input_tokens,
            "output_tokens": job.output_tokens,
            "cost_usd": str(job.cost_usd) if job.cost_usd is not None else None,
        },
    )
    return result


def reject_extraction(
    db: Session,
    *,
    job_id: str,
    user_id: str | None,
    user_email: str | None = None,
    reason: str | None = None,
) -> ExtractionJob:
    job = db.get(ExtractionJob, job_id)
    if job is None:
        raise ExtractionError(f"ExtractionJob {job_id} not found")
    job.status = ExtractionStatus.REJECTED
    job.error_message = reason
    job.confirmed_at = datetime.now(timezone.utc)
    job.confirmed_by = user_id
    db.flush()
    AuditService(db).log(
        user_email=user_email or "system",
        action="extraction.rejected",
        workspace_id=job.workspace_id,
        user_id=user_id,
        resource_type="extraction_job",
        resource_id=job.id,
        details={"reason": reason},
    )
    return job


# ── apply per hint ───────────────────────────────────────────────────────────

def _apply_payload(
    db: Session,
    job: ExtractionJob,
    payload: dict,
    *,
    user_id: str | None,
    user_email: str | None,
    institution_short_name: str | None = None,
    manual_mappings: dict[str, str] | None = None,
    manual_prices: dict[str, float] | None = None,
    manual_modes: dict[str, str] | None = None,
) -> ExtractionApplyResult:
    hint = job.source_hint
    # Spec 48 — bulk path: BROKER_POSITION + snapshot scope + no specific
    # pendency → resolve N pendencies in the snapshot.
    if (
        hint == ExtractionSourceHint.BROKER_POSITION
        and job.snapshot_id
        and not job.pendency_id
    ):
        return _apply_bulk_to_snapshot(
            db, job, payload,
            user_id=user_id, user_email=user_email,
            institution_short_name=institution_short_name,
            manual_mappings=manual_mappings,
            manual_prices=manual_prices,
            manual_modes=manual_modes,
        )
    if hint == ExtractionSourceHint.SCREENSHOT_PRICE:
        return _apply_screenshot_price(db, job, payload, user_id=user_id)
    if hint == ExtractionSourceHint.BROKER_POSITION:
        return _apply_broker_position(db, job, payload, user_id=user_id)
    # Spec 58 Stage 4 — BROKER_INCOME (proventos) bulk path.
    if (
        hint == ExtractionSourceHint.BROKER_INCOME
        and job.snapshot_id
        and not job.pendency_id
    ):
        return _apply_bulk_income_to_snapshot(
            db, job, payload, user_id=user_id, user_email=user_email,
        )
    # Other hints are scaffold-only in V1.
    return ExtractionApplyResult(
        applied_count=0, skipped_count=0,
        errors=[f"applying hint {hint.value} not yet implemented in V1"],
    )


def _apply_screenshot_price(
    db: Session, job: ExtractionJob, payload: dict, *, user_id: str | None,
) -> ExtractionApplyResult:
    """SCREENSHOT_PRICE → update Asset.current_price for the matching ticker
    in the same workspace. If the pendency carries an asset_id, prefer that
    over the LLM's `ticker` (the user uploaded the screenshot _for_ that
    asset)."""
    target_asset: Asset | None = None
    pendency = db.get(SnapshotPendency, job.pendency_id) if job.pendency_id else None
    if pendency:
        target_asset = db.get(Asset, pendency.asset_id)
    elif job.asset_id:
        target_asset = db.get(Asset, job.asset_id)
    else:
        ticker = payload.get("ticker")
        if ticker:
            target_asset = (
                db.query(Asset)
                .filter(
                    Asset.workspace_id == job.workspace_id,
                    Asset.ticker == ticker,
                )
                .first()
            )

    if target_asset is None:
        return ExtractionApplyResult(
            applied_count=0, skipped_count=1,
            errors=[
                f"no Asset matches ticker={payload.get('ticker')!r} in workspace",
            ],
        )

    price = payload.get("price")
    if price is None:
        return ExtractionApplyResult(
            applied_count=0, skipped_count=1, errors=["missing price"],
        )

    target_asset.current_price = Decimal(str(price))
    target_asset.price_updated_at = datetime.now(timezone.utc)
    db.flush()
    return ExtractionApplyResult(applied_count=1, skipped_count=0, errors=[])


def _apply_broker_position(
    db: Session, job: ExtractionJob, payload: dict, *, user_id: str | None,
) -> ExtractionApplyResult:
    """BROKER_POSITION → update Asset.current_price for each line we can
    resolve to a workspace asset. Unmatched lines are reported.

    Assets with an automated price source (BRAPI/FINNHUB/COINBASE/TESOURO)
    are skipped — those prices are owned by their provider, not by
    user-uploaded extracts.
    """
    from numis_geek.services.price_freshness import AUTOMATED_SOURCES

    positions = payload.get("positions") or []
    applied = 0
    skipped = 0
    errors: list[str] = []
    now = datetime.now(timezone.utc)
    for pos in positions:
        ticker = pos.get("ticker_normalized") or pos.get("ticker_raw")
        if not ticker:
            skipped += 1
            errors.append("position with no ticker")
            continue
        asset = _resolve_asset_by_ticker_or_name(
            db, job.workspace_id, ticker, institution_id=job.institution_id,
        )
        if asset is None:
            skipped += 1
            errors.append(f"ticker {ticker!r} not found")
            continue
        if asset.price_source in AUTOMATED_SOURCES:
            skipped += 1
            errors.append(
                f"ticker {ticker!r}: preço gerenciado por {asset.price_source.value} "
                "— extrato ignorado (use refresh da API)."
            )
            continue
        price = pos.get("unit_price")
        if price is None:
            skipped += 1
            continue
        asset.current_price = Decimal(str(price))
        asset.price_updated_at = now
        applied += 1
    db.flush()
    return ExtractionApplyResult(
        applied_count=applied, skipped_count=skipped, errors=errors,
    )


# ── Spec 48 — bulk apply to a snapshot ──────────────────────────────────────


def _alnum_lower(s: str | None) -> str:
    """Aggressive normalization for substring matching: keep only
    alphanumeric chars, lowercased. Collapses punctuation/whitespace
    differences (`S/T` vs `-ST`, `U.S.` vs `US`, `A(acc)USD` vs `Aacc`)."""
    if not s:
        return ""
    return "".join(c.lower() for c in s if c.isalnum())


# Date patterns for maturity-based matching across asset name/ticker and
# extracted ticker_raw. Brokers use wildly different formats:
#   - ISO: '2034-05-16'
#   - BR: '16/05/2034' or '16/05/34'
#   - US: '05/16/2034' or '05/16/34'
_DATE_RE_ISO = re.compile(r"(?<!\d)(\d{4})[-/](\d{1,2})[-/](\d{1,2})(?!\d)")
_DATE_RE_DMY4 = re.compile(r"(?<!\d)(\d{1,2})[-/](\d{1,2})[-/](\d{4})(?!\d)")
_DATE_RE_DMY2 = re.compile(r"(?<!\d)(\d{1,2})[-/](\d{1,2})[-/](\d{2})(?!\d)")


def _extract_dates(s: str | None) -> set[str]:
    """Return all ISO YYYY-MM-DD dates extractable from `s`. For 2-digit
    years with a DD/MM-vs-MM/DD ambiguity (e.g. '11/04/29'), returns BOTH
    interpretations when both are valid month-day combos. Year pivot: YY
    < 70 → 20YY, else 19YY (matches industry convention for maturity
    dates that are always in the future or recent past)."""
    if not s:
        return set()
    out: set[str] = set()
    for m in _DATE_RE_ISO.finditer(s):
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mo <= 12 and 1 <= d <= 31:
            out.add(f"{y:04d}-{mo:02d}-{d:02d}")
    for m in _DATE_RE_DMY4.finditer(s):
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mo <= 12 and 1 <= d <= 31:
            out.add(f"{y:04d}-{mo:02d}-{d:02d}")
    for m in _DATE_RE_DMY2.finditer(s):
        a, b, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
        full_year = 2000 + yy if yy < 70 else 1900 + yy
        # DD/MM/YY interpretation
        if 1 <= b <= 12 and 1 <= a <= 31:
            out.add(f"{full_year:04d}-{b:02d}-{a:02d}")
        # MM/DD/YY interpretation
        if 1 <= a <= 12 and 1 <= b <= 31:
            out.add(f"{full_year:04d}-{a:02d}-{b:02d}")
    return out


def _resolve_asset_by_ticker_or_name(
    db: Session, workspace_id: str, candidate: str,
    *,
    institution_id: str | None = None,
) -> Asset | None:
    """Match an extracted line to a workspace Asset.

    1. Case-insensitive ticker match (with strip) — LLM output isn't
       guaranteed to be uppercase or trimmed; DB tickers may also have
       mixed case for funds/CDBs.
    2. Case-insensitive name match — handles funds, fixed income, and other
       assets registered without a canonical exchange ticker (e.g. the
       `Fundo Verde BTG` case where Asset.ticker carries the fund name)
    3. Normalized substring match against `Asset.name` AND `Asset.ticker`,
       either direction — last-resort fuzz. Normalization strips
       punctuation/whitespace, so `Franklin U.S. Dollar` (ticker) matches
       `Franklin U.S. Dollar-ST MMF Adv` (extract). Only returns when
       exactly one asset matches, to avoid spurious resolutions.
    4. Date-based match for fixed income: extracts ISO/BR/US date formats
       from candidate and from each asset's name+ticker. If exactly one
       asset shares a maturity date with the candidate, return it. Handles
       `JPMorgan Chase 2033-09-14` (extract) ↔ `JPM 5.717 14/09/33`
       (asset) — same date, different formats.

    Spec 58 — when `institution_id` is set, restricts the candidate pool
    to assets whose account belongs to that FI. Shrinks the matcher's
    universe → less ambiguity in steps 3 and 4.
    """
    if not candidate:
        return None
    needle = candidate.strip().lower()
    if not needle:
        return None

    def _scope(q):
        if institution_id is None:
            return q
        return (
            q.join(Account, Asset.account_id == Account.id, isouter=False)
             .filter(Account.financial_institution_id == institution_id)
        )

    base = db.query(Asset).filter(
        Asset.workspace_id == workspace_id,
        Asset.is_active == True,  # noqa: E712
    )
    # Step 1 — ticker, case-insensitive + trimmed on both sides.
    hit = _scope(base.filter(
        func.lower(func.trim(Asset.ticker)) == needle,
    )).first()
    if hit is not None:
        return hit
    # Step 2 — exact name (case-insensitive).
    all_assets = _scope(base).all()
    hit = next(
        (a for a in all_assets if a.name and a.name.lower() == needle),
        None,
    )
    if hit is not None:
        return hit
    # Step 3 — normalized substring against name AND ticker.
    needle_norm = _alnum_lower(candidate)
    if len(needle_norm) < 4:
        return None
    candidates = all_assets
    matches: list[Asset] = []
    for a in candidates:
        for asset_str in (a.name, a.ticker):
            asset_norm = _alnum_lower(asset_str)
            if len(asset_norm) < 4:
                continue
            if asset_norm in needle_norm or needle_norm in asset_norm:
                matches.append(a)
                break  # one hit per asset is enough
    if len(matches) == 1:
        return matches[0]
    # Step 4 — date overlap. Fixed income often shares only a maturity
    # date across naming variants (Asset.ticker='JPM 5.717 14/09/33',
    # extract='JPMorgan Chase 2033-09-14'). Run only when Step 3 didn't
    # already produce a single hit.
    candidate_dates = _extract_dates(candidate)
    if not candidate_dates:
        return None
    date_matches: list[Asset] = []
    for a in candidates:
        asset_dates = _extract_dates(a.ticker) | _extract_dates(a.name)
        if candidate_dates & asset_dates:
            date_matches.append(a)
    return date_matches[0] if len(date_matches) == 1 else None


def _institution_short_name_for_asset(db: Session, asset: Asset) -> str | None:
    if not asset.account_id:
        return None
    acc = db.get(Account, asset.account_id)
    if acc is None or not acc.financial_institution_id:
        return None
    fi = db.get(FinancialInstitution, acc.financial_institution_id)
    return fi.short_name if fi else None


def _classify_bulk_extract(
    db: Session, job: ExtractionJob, payload: dict,
    *,
    institution_short_name: str | None = None,
    manual_mappings: dict[str, str] | None = None,
    manual_prices: dict[str, float] | None = None,
    manual_modes: dict[str, str] | None = None,
) -> tuple[
    list[dict],  # apply_plan: items the writer must process (incl. pendency, asset, price, mode)
    BulkApplyDetail,
    list[str],   # classification-time errors (e.g. missing ticker, snapshot not found)
]:
    # Spec 58 — job-scoped FI overrides body-provided short_name. When
    # set, we ALSO restrict the candidate Asset pool to accounts at this
    # FI (not just filter the matched set at the end). Smaller pool →
    # less ambiguity in the substring/date matcher.
    scoped_fi_id: str | None = job.institution_id
    if scoped_fi_id:
        fi = db.get(FinancialInstitution, scoped_fi_id)
        if fi is None:
            raise ExtractionError(f"institution {scoped_fi_id} not found")
        institution_short_name = fi.short_name
    """Spec 57 follow-up — pure classification, no writes.

    Returns the prospective plan PLUS the categorized buckets the UI shows.
    `_apply_bulk_to_snapshot` calls this and then performs writes for each
    item in `apply_plan`. The preview endpoint calls this and discards
    `apply_plan` (or reuses it to show predicted new/previous prices).
    """
    snap = db.get(PortfolioSnapshot, job.snapshot_id) if job.snapshot_id else None
    if snap is None:
        return [], BulkApplyDetail(
            applied=[], matched_no_pendency=[], orphan=[],
            pendency_not_in_extract=[], auto_skipped=[],
        ), [f"snapshot {job.snapshot_id} not found"]

    # Bulk extract only resolves pendencies whose action is "user edits the
    # price" (EDIT_PRICE/UPLOAD_FILE). RETRY_API pendencies belong to assets
    # with an automated price source — their fix is re-fetching the provider,
    # not overwriting the price from a brokerage extract.
    from numis_geek.models.portfolio_snapshot import PendencyAction
    from numis_geek.services.price_freshness import AUTOMATED_SOURCES

    open_pendencies = (
        db.query(SnapshotPendency)
        .filter(
            SnapshotPendency.snapshot_id == snap.id,
            SnapshotPendency.resolved_at.is_(None),
            SnapshotPendency.action_type != PendencyAction.RETRY_API,
        )
        .all()
    )
    pendency_by_asset = {p.asset_id: p for p in open_pendencies}

    positions = payload.get("positions") or []
    apply_plan: list[dict] = []
    applied_preview: list[dict] = []
    matched_no_pendency: list[dict] = []
    orphan: list[dict] = []
    auto_skipped: list[dict] = []
    matched_asset_ids: set[str] = set()
    errors: list[str] = []

    overrides = dict(manual_mappings or {})
    price_overrides = dict(manual_prices or {})
    mode_overrides = dict(manual_modes or {})

    for pos in positions:
        ticker = pos.get("ticker_normalized") or pos.get("ticker_raw")
        ticker_raw_key = pos.get("ticker_raw") or ""
        manual_price = (
            price_overrides.get(ticker_raw_key)
            or price_overrides.get(ticker)
        )
        # Sanity check qty × unit_price ≈ market_value (2026-07-04). Bug
        # sistemático em extração Avenue: Claude misturava custo médio com
        # preço atual, dando erros de 10-50% no total. Quando LLM devolve
        # os três e a conta não bate, o market_value costuma ser o número
        # certo (é o mais destacado no extrato) — reescreve unit_price a
        # partir dele pra preservar o total que o user vê.
        llm_qty = pos.get("quantity")
        llm_unit = pos.get("unit_price")
        llm_mv = pos.get("market_value")
        if (
            manual_price is None
            and llm_qty is not None
            and llm_unit is not None
            and llm_mv is not None
        ):
            try:
                q = float(llm_qty)
                p = float(llm_unit)
                mv = float(llm_mv)
                computed = q * p
                # % de face value: prompt v4 Rule 5 diz Treasuries/bonds
                # cotados em % têm mv = qty × price / 100 (ex.: 10000 face
                # × 99.943 = 9994.30). Se a conta bate com /100, aceita
                # sem reconciliar. Audit 2026-07-05 pegou reconciler
                # destruindo unit_price de Treasuries: 99.943 → 0.99943.
                if q > 0 and mv > 0 and abs(computed / 100 - mv) / mv <= 0.01:
                    pass  # % of face, invariante OK
                elif q > 0 and mv > 0 and abs(computed - mv) / mv > 0.01:
                    reconciled_unit = mv / q
                    if reconciled_unit <= 0:
                        errors.append(
                            f"{ticker}: reconciled unit_price seria {reconciled_unit:.4f} (≤0); mantendo LLM value e sinalizando review"
                        )
                    else:
                        errors.append(
                            f"{ticker}: qty×price ({computed:.2f}) ≠ mv ({mv:.2f}); "
                            f"unit_price recomputed from mv/qty = {reconciled_unit:.4f}"
                        )
                        llm_unit = reconciled_unit
                elif q == 0 and mv > 0:
                    errors.append(
                        f"{ticker}: qty=0 mas mv={mv:.2f}; verificar se posição zerada ou mv errado"
                    )
            except (TypeError, ValueError, ZeroDivisionError):
                pass
        # Fallback: se LLM omitiu unit_price mas mandou qty + mv, deriva.
        # Não usa llm_mv direto como unit_price (bug pré-existente: setava
        # asset.current_price = valor TOTAL da posição pra STOCK/ETF).
        if manual_price is None and llm_unit is None and llm_qty is not None and llm_mv is not None:
            try:
                q = float(llm_qty)
                if q > 0:
                    llm_unit = float(llm_mv) / q
                    errors.append(
                        f"{ticker}: LLM omitiu unit_price; derivado de mv/qty = {llm_unit:.4f}"
                    )
            except (TypeError, ValueError, ZeroDivisionError):
                pass
        # Fallback llm_mv REMOVIDO (2026-07-05): usar mv como unit_price
        # setava Asset.current_price = valor TOTAL da posição pra
        # STOCK/ETF (ex.: AAPL 100 shares × $150 = mv 15000 virava
        # current_price=15000). Se falta unit_price real, agora é
        # derivado de mv/qty acima antes de chegar aqui.
        unit_price_raw = manual_price or llm_unit
        if not ticker:
            errors.append("position with no ticker")
            continue
        manual_pendency_id = overrides.get(ticker_raw_key) or overrides.get(ticker)
        asset: Asset | None = None
        if manual_pendency_id:
            forced_pen = db.get(SnapshotPendency, manual_pendency_id)
            if forced_pen and forced_pen.snapshot_id == snap.id:
                asset = db.get(Asset, forced_pen.asset_id)
        if asset is None:
            asset = _resolve_asset_by_ticker_or_name(
                db, job.workspace_id, ticker, institution_id=scoped_fi_id,
            )
        if asset is None:
            # Currency only from LLM payload for orphan (no asset to look up).
            orphan.append({
                "ticker": ticker,
                "unit_price": str(unit_price_raw) if unit_price_raw is not None else None,
                "currency": pos.get("currency") or None,
            })
            continue

        asset_fi = _institution_short_name_for_asset(db, asset)
        asset_currency = asset.currency.value if asset.currency else None

        # Auto-priced asset: extract is informational only. Surface in
        # auto_skipped so the UI can show "recognized, no action needed",
        # but never feed it into the apply plan (provider owns the price).
        if asset.price_source in AUTOMATED_SOURCES and not manual_pendency_id:
            auto_skipped.append({
                "asset_id": asset.id,
                "ticker": ticker,
                "asset_name": asset.name,
                "currency": asset_currency,
                "institution_short_name": asset_fi,
                "price_source": asset.price_source.value,
                "unit_price": str(unit_price_raw) if unit_price_raw is not None else None,
            })
            continue

        if (
            institution_short_name
            and asset_fi != institution_short_name
            and not manual_pendency_id
        ):
            # Different FI than the one scoped: silently out of scope.
            continue

        matched_asset_ids.add(asset.id)
        pendency = pendency_by_asset.get(asset.id)
        if pendency is None:
            matched_no_pendency.append({
                "asset_id": asset.id,
                "ticker": ticker,
                "asset_name": asset.name,
                "currency": asset_currency,
                "institution_short_name": asset_fi,
                "unit_price": str(unit_price_raw) if unit_price_raw is not None else None,
            })
            continue
        if unit_price_raw is None:
            errors.append(
                f"{ticker!r}: sem preço unitário no extrato "
                "(use o input ao lado do dropdown pra informar manualmente)."
            )
            continue

        effective_mode = (
            mode_overrides.get(ticker_raw_key)
            or mode_overrides.get(ticker)
        )
        new_price = Decimal(str(unit_price_raw))
        apply_plan.append({
            "pendency_id": pendency.id,
            "asset_id": asset.id,
            "ticker": ticker,
            "asset_name": asset.name,
            "currency": asset_currency,
            "institution_short_name": asset_fi,
            "new_price": new_price,
            "previous_price": asset.current_price,
            "effective_mode": effective_mode,
            "ticker_raw_key": ticker_raw_key,
        })
        applied_preview.append({
            "pendency_id": pendency.id,
            "asset_id": asset.id,
            "ticker": ticker,
            "asset_name": asset.name,
            "currency": asset_currency,
            "institution_short_name": asset_fi,
            "new_price": str(new_price),
            "previous_price": (
                str(asset.current_price) if asset.current_price is not None else None
            ),
        })

    # Pendencies still open after this pass — scoped to the FI when set.
    not_in_extract: list[dict] = []
    for p in open_pendencies:
        if p.asset_id in matched_asset_ids:
            continue
        asset = db.get(Asset, p.asset_id)
        if asset is None:
            continue
        asset_fi = _institution_short_name_for_asset(db, asset)
        if institution_short_name and asset_fi != institution_short_name:
            continue
        not_in_extract.append({
            "pendency_id": p.id,
            "asset_id": asset.id,
            "ticker": asset.ticker,
            "asset_name": asset.name,
            "currency": asset.currency.value if asset.currency else None,
            "institution_short_name": asset_fi,
        })

    return apply_plan, BulkApplyDetail(
        applied=applied_preview,
        matched_no_pendency=matched_no_pendency,
        orphan=orphan,
        pendency_not_in_extract=not_in_extract,
        auto_skipped=auto_skipped,
    ), errors


def preview_bulk_extract(
    db: Session, *, job_id: str,
    institution_short_name: str | None = None,
    manual_mappings: dict[str, str] | None = None,
    manual_prices: dict[str, float] | None = None,
    manual_modes: dict[str, str] | None = None,
) -> ExtractionApplyResult:
    """Read-only classification of an EXTRACTED bulk job. Returns the same
    `BulkApplyDetail` shape the UI gets after Apply, but no writes happen.
    Useful for the review modal so it shows exactly what Apply will do."""
    job = db.get(ExtractionJob, job_id)
    if job is None:
        raise ExtractionError(f"Job {job_id} not found")
    if job.snapshot_id is None:
        raise ExtractionError("Preview only available for bulk jobs (snapshot_id set)")
    payload = job.extracted_json or {}
    apply_plan, detail, errors = _classify_bulk_extract(
        db, job, payload,
        institution_short_name=institution_short_name,
        manual_mappings=manual_mappings,
        manual_prices=manual_prices,
        manual_modes=manual_modes,
    )
    _ = apply_plan  # discarded — preview is read-only
    return ExtractionApplyResult(
        applied_count=len(detail.applied),
        skipped_count=len(detail.orphan) + len(detail.matched_no_pendency),
        errors=errors,
        bulk_detail=detail,
    )


def _apply_bulk_to_snapshot(
    db: Session, job: ExtractionJob, payload: dict,
    *,
    user_id: str | None,
    user_email: str | None,
    institution_short_name: str | None = None,
    manual_mappings: dict[str, str] | None = None,
    manual_prices: dict[str, float] | None = None,
    manual_modes: dict[str, str] | None = None,
) -> ExtractionApplyResult:
    """Bulk path (Spec 48): match positions[] by ticker to open pendencies
    in the snapshot. Each match calls resolve_pendency (which updates
    Asset.current_price + PortfolioSnapshotItem + recomputes totals) with
    the bulk attachment as the resolution attachment.

    Returns categorized lists in BulkApplyDetail so the UI can render
    the review/result panel without re-querying.
    """
    apply_plan, detail, errors = _classify_bulk_extract(
        db, job, payload,
        institution_short_name=institution_short_name,
        manual_mappings=manual_mappings,
        manual_prices=manual_prices,
        manual_modes=manual_modes,
    )
    if not apply_plan and not detail.applied:
        # Snapshot-not-found case: detail is empty + error already set.
        return ExtractionApplyResult(
            applied_count=0, skipped_count=0, errors=errors, bulk_detail=detail,
        )

    from numis_geek.services import snapshot as snapshot_service

    actually_applied: list[dict] = []
    for item in apply_plan:
        note = (
            f"bulk extract (job {job.id})"
            + (f" · FI={institution_short_name}" if institution_short_name else "")
            + (f" · mode={item['effective_mode']}" if item["effective_mode"] else "")
        )
        try:
            snapshot_service.resolve_pendency(
                db,
                pendency_id=item["pendency_id"],
                user_id=user_id,
                user_email=user_email,
                new_price=item["new_price"],
                value_mode=item["effective_mode"],
                file_id=job.attachment_id,
                note=note,
            )
        except ValueError as e:
            errors.append(f"pendency {item['pendency_id']}: {e}")
            continue
        actually_applied.append({
            "pendency_id": item["pendency_id"],
            "asset_id": item["asset_id"],
            "ticker": item["ticker"],
            "asset_name": item["asset_name"],
            "currency": item.get("currency"),
            "institution_short_name": item["institution_short_name"],
            "new_price": str(item["new_price"]),
            "previous_price": (
                str(item["previous_price"]) if item["previous_price"] is not None else None
            ),
        })

    db.flush()
    # Overwrite the prospective `applied` with what actually went through.
    detail.applied = actually_applied
    return ExtractionApplyResult(
        applied_count=len(actually_applied),
        skipped_count=len(detail.orphan) + len(detail.matched_no_pendency),
        errors=errors,
        bulk_detail=detail,
    )


# ── Spec 58 Stage 4 — BROKER_INCOME apply (Distribution rows) ───────────────


_DIST_TYPE_BY_PAYLOAD: dict[str, DistributionType] = {
    "DIVIDEND": DistributionType.DIVIDEND,
    "INTEREST": DistributionType.INTEREST,
    "JCP": DistributionType.JCP,
    "SECURITIES_LENDING": DistributionType.SECURITIES_LENDING,
}


def _income_period_bounds(snap: PortfolioSnapshot) -> tuple[date, date]:
    """Janela mensal coberta pelo snapshot — primeiro dia do mês até
    period_end_date inclusive. Eventos com event_date fora desse range
    não pertencem a esse fechamento."""
    period_end = snap.period_end_date
    period_start = period_end.replace(day=1)
    return period_start, period_end


def _external_id_for_income(
    *, fi_short_name: str, event_date_iso: str, ticker_raw: str | None,
    type_str: str, gross_amount: float,
) -> str:
    """Content-addressed id for idempotency. Re-uploading the same CSV
    (or a re-export with identical rows) won't duplicate Distribution
    rows — the (external_source, external_id) index dedups."""
    safe_fi = (fi_short_name or "unknown").strip().lower().replace(" ", "_")
    safe_ticker = (ticker_raw or "_").strip()
    gross_str = f"{float(gross_amount):.2f}"
    return f"{safe_fi}:{event_date_iso}:{safe_ticker}:{type_str}:{gross_str}"


def _classify_bulk_income(
    db: Session, job: ExtractionJob, payload: dict,
) -> tuple[
    list[dict],   # apply_plan
    list[dict],   # preview rows (matched)
    list[dict],   # orphan: events whose ticker didn't resolve to an Asset
    list[dict],   # duplicate: external_id already present in DB
    list[str],    # errors
]:
    """Read-only classification of a BROKER_INCOME payload. Resolves
    each event to an Asset (when applicable) and looks up existing
    Distribution rows by external_id to flag duplicates."""
    errors: list[str] = []
    apply_plan: list[dict] = []
    preview: list[dict] = []
    orphan: list[dict] = []
    duplicate: list[dict] = []

    if not job.institution_id:
        errors.append("BROKER_INCOME job must be scoped to an institution")
        return apply_plan, preview, orphan, duplicate, errors

    fi = db.get(FinancialInstitution, job.institution_id)
    if fi is None:
        errors.append(f"institution {job.institution_id} not found")
        return apply_plan, preview, orphan, duplicate, errors

    # Garante que só eventos do período do snapshot sejam aplicados —
    # extratos costumam vir multi-mês (ex.: Mar a Mai num único XLSX).
    period_start: date | None = None
    period_end: date | None = None
    if job.snapshot_id:
        snap = db.get(PortfolioSnapshot, job.snapshot_id)
        if snap:
            period_start, period_end = _income_period_bounds(snap)

    events = payload.get("events") or []
    out_of_period = 0

    for ev in events:
        type_str = (ev.get("type") or "").upper()
        if type_str not in _DIST_TYPE_BY_PAYLOAD:
            errors.append(f"unknown distribution type: {type_str!r}")
            continue
        dist_type = _DIST_TYPE_BY_PAYLOAD[type_str]
        event_date_iso = ev.get("event_date")
        gross = ev.get("gross_amount")
        net = ev.get("net_amount")
        tax = ev.get("tax_amount")
        if event_date_iso is None or gross is None or net is None:
            errors.append(f"event missing required fields: {ev!r}")
            continue
        try:
            event_d = datetime.strptime(event_date_iso, "%Y-%m-%d").date()
        except (TypeError, ValueError):
            errors.append(f"invalid event_date {event_date_iso!r}")
            continue
        if period_start and period_end:
            if event_d < period_start or event_d > period_end:
                out_of_period += 1
                continue

        ticker = ev.get("ticker_raw")
        asset: Asset | None = None
        if ticker:
            asset = _resolve_asset_by_ticker_or_name(
                db, job.workspace_id, ticker, institution_id=job.institution_id,
            )

        currency_str = (ev.get("currency") or "USD").upper()
        try:
            currency_enum = Currency(currency_str)
        except ValueError:
            errors.append(f"unknown currency {currency_str!r}")
            continue

        ext_id = _external_id_for_income(
            fi_short_name=fi.short_name,
            event_date_iso=event_date_iso,
            ticker_raw=ticker,
            type_str=type_str,
            gross_amount=float(gross),
        )

        existing = (
            db.query(Distribution)
            .filter(
                Distribution.workspace_id == job.workspace_id,
                Distribution.external_source == ExternalSource.MANUAL_CSV,
                Distribution.external_id == ext_id,
            )
            .first()
        )

        row = {
            "external_id": ext_id,
            "event_date": event_date_iso,
            "ticker": ticker,
            "asset_id": asset.id if asset else None,
            "asset_name": asset.name if asset else None,
            "type": type_str,
            "gross_amount": str(Decimal(str(gross))),
            "tax_amount": str(Decimal(str(tax))) if tax is not None else None,
            "net_amount": str(Decimal(str(net))),
            "currency": currency_str,
            "institution_short_name": fi.short_name,
        }

        if existing is not None:
            row["distribution_id"] = existing.id
            duplicate.append(row)
            continue

        # Orphan = ticker present but no asset matched. Lending (ticker=None)
        # is allowed without an asset (Distribution.asset_id is nullable).
        if ticker and asset is None:
            orphan.append(row)
            continue

        apply_plan.append({
            **row,
            "_currency_enum": currency_enum,
            "_dist_type": dist_type,
            "_event_d": event_d,
            "_gross": Decimal(str(gross)),
            "_net": Decimal(str(net)),
            "_tax": Decimal(str(tax)) if tax is not None else None,
        })
        preview.append(row)

    if out_of_period > 0 and period_start and period_end:
        errors.append(
            f"{out_of_period} evento(s) fora do período "
            f"{period_start.isoformat()}–{period_end.isoformat()} ignorado(s)"
        )

    return apply_plan, preview, orphan, duplicate, errors


_OPTION_SIDE_BY_PAYLOAD: dict[str, AssetMovementType] = {
    "SELL_OPEN": AssetMovementType.SELL_OPEN,
    "BUY_TO_CLOSE": AssetMovementType.BUY_TO_CLOSE,
}


def _external_id_for_option_event(
    *, fi_short_name: str, event_date_iso: str, option_ticker_raw: str,
    side: str, gross_amount: float,
) -> str:
    """Content-addressed id for option premium movements. Same shape as
    distribution external_id pra idempotência: re-upload do mesmo extrato
    não duplica AssetMovement."""
    norm_ticker = option_ticker_raw.strip().upper()
    return (
        f"{fi_short_name}|{event_date_iso}|{norm_ticker}|"
        f"OPTION:{side}|{gross_amount:.2f}"
    )


def _classify_bulk_option_events(
    db: Session, job: ExtractionJob, payload: dict,
) -> tuple[
    list[dict],   # apply_plan
    list[dict],   # preview
    list[dict],   # orphan: option_ticker não bate com nenhum Asset OPTION
    list[dict],   # duplicate: external_id já existe em AssetMovement
    list[str],    # errors
]:
    """Classifica option_events do payload do BROKER_INCOME. Roteia pra
    AssetMovement (não Distribution) — single source of truth pra opções."""
    errors: list[str] = []
    apply_plan: list[dict] = []
    preview: list[dict] = []
    orphan: list[dict] = []
    duplicate: list[dict] = []

    if not job.institution_id:
        return apply_plan, preview, orphan, duplicate, errors
    fi = db.get(FinancialInstitution, job.institution_id)
    if fi is None:
        return apply_plan, preview, orphan, duplicate, errors

    period_start: date | None = None
    period_end: date | None = None
    if job.snapshot_id:
        snap = db.get(PortfolioSnapshot, job.snapshot_id)
        if snap:
            period_start, period_end = _income_period_bounds(snap)

    option_events = payload.get("option_events") or []
    out_of_period = 0

    for ev in option_events:
        side_str = (ev.get("side") or "").upper()
        if side_str not in _OPTION_SIDE_BY_PAYLOAD:
            errors.append(f"unknown option side: {side_str!r}")
            continue
        side_enum = _OPTION_SIDE_BY_PAYLOAD[side_str]
        event_date_iso = ev.get("event_date")
        ticker = ev.get("option_ticker_raw")
        qty = ev.get("quantity")
        unit_price = ev.get("unit_price")
        gross = ev.get("gross_amount")
        if event_date_iso is None or ticker is None or qty is None or gross is None:
            errors.append(f"option event missing required fields: {ev!r}")
            continue
        try:
            event_d = datetime.strptime(event_date_iso, "%Y-%m-%d").date()
        except (TypeError, ValueError):
            errors.append(f"invalid event_date {event_date_iso!r}")
            continue
        if period_start and period_end:
            if event_d < period_start or event_d > period_end:
                out_of_period += 1
                continue

        # Lookup do Asset OPTION pelo ticker exato (workspace + class=OPTION).
        asset = (
            db.query(Asset)
            .filter(
                Asset.workspace_id == job.workspace_id,
                Asset.asset_class == AssetClass.OPTION,
                Asset.ticker == ticker.strip().upper(),
            )
            .first()
        )

        currency_str = (ev.get("currency") or "BRL").upper()
        try:
            currency_enum = Currency(currency_str)
        except ValueError:
            errors.append(f"unknown currency {currency_str!r}")
            continue

        net = ev.get("net_amount")
        fee = ev.get("fee")
        net_decimal = (
            Decimal(str(net)) if net is not None else Decimal(str(gross))
        )

        ext_id = _external_id_for_option_event(
            fi_short_name=fi.short_name,
            event_date_iso=event_date_iso,
            option_ticker_raw=ticker,
            side=side_str,
            gross_amount=float(gross),
        )
        # Dedup em 2 camadas:
        # 1) external_id MANUAL_CSV — catches re-upload do mesmo extrato.
        # 2) (asset, type, date, gross) — catches lançamento manual feito
        #    antes do upload (sem external_id). Sem isso, o extrato
        #    duplicaria movements que o user já registrou na mão.
        existing = None
        if asset is not None:
            existing = (
                db.query(AssetMovement)
                .filter(
                    AssetMovement.workspace_id == job.workspace_id,
                    AssetMovement.asset_id == asset.id,
                    AssetMovement.type == side_enum,
                    AssetMovement.event_date == event_d,
                    AssetMovement.gross_amount == Decimal(str(gross)),
                    AssetMovement.is_active.is_(True),
                )
                .first()
            )
        if existing is None:
            existing = (
                db.query(AssetMovement)
                .filter(
                    AssetMovement.workspace_id == job.workspace_id,
                    AssetMovement.external_source == ExternalSource.MANUAL_CSV,
                    AssetMovement.external_id == ext_id,
                )
                .first()
            )

        row = {
            "external_id": ext_id,
            "event_date": event_date_iso,
            "option_ticker": ticker,
            "side": side_str,
            "asset_id": asset.id if asset else None,
            "asset_name": asset.name if asset else None,
            "quantity": str(Decimal(str(qty))),
            "unit_price": str(Decimal(str(unit_price))) if unit_price is not None else None,
            "gross_amount": str(Decimal(str(gross))),
            "fee": str(Decimal(str(fee))) if fee is not None else None,
            "net_amount": str(net_decimal),
            "currency": currency_str,
            "institution_short_name": fi.short_name,
            "row_kind": "option_event",
        }

        if existing is not None:
            row["movement_id"] = existing.id
            duplicate.append(row)
            continue
        if asset is None:
            orphan.append(row)
            continue

        apply_plan.append({
            **row,
            "_asset": asset,
            "_side_enum": side_enum,
            "_currency_enum": currency_enum,
            "_event_d": event_d,
            "_qty": Decimal(str(qty)),
            "_unit_price": Decimal(str(unit_price)) if unit_price is not None else None,
            "_gross": Decimal(str(gross)),
            "_fee": Decimal(str(fee)) if fee is not None else None,
            "_net": net_decimal,
        })
        preview.append(row)

    if out_of_period > 0 and period_start and period_end:
        errors.append(
            f"{out_of_period} option_event(s) fora do período "
            f"{period_start.isoformat()}–{period_end.isoformat()} ignorado(s)"
        )

    return apply_plan, preview, orphan, duplicate, errors


def _apply_bulk_income_to_snapshot(
    db: Session, job: ExtractionJob, payload: dict,
    *,
    user_id: str | None,
    user_email: str | None,
) -> ExtractionApplyResult:
    """BROKER_INCOME path (Spec 58 Stage 4): create Distribution rows
    from a proventos extract. Idempotent via (external_source, external_id).

    Linhas de prêmio de opção (payload.option_events) viram AssetMovement
    SELL_OPEN/BUY_TO_CLOSE em vez de Distribution — o sintético
    OPTION_PREMIUM é derivado disso pelo services/proventos.py."""
    apply_plan, preview_rows, orphan_rows, duplicate_rows, errors = (
        _classify_bulk_income(db, job, payload)
    )
    (
        opt_plan, opt_preview, opt_orphan, opt_duplicate, opt_errors,
    ) = _classify_bulk_option_events(db, job, payload)
    errors = errors + opt_errors

    created: list[dict] = []
    fi_id = job.institution_id
    now = datetime.now(timezone.utc)
    for item in apply_plan:
        # Convenção Distribution: fx_rate é MULTIPLICADOR pra BRL.
        # BRL native → 1.0; USD native → PTAX da event_date.
        # Hardcode antigo de 1.0 fazia $X net virar R$ X na KPI (sub-conta ~5x).
        is_usd = item["_currency_enum"] == Currency.USD
        dist_fx = (
            resolve_fx_rate(db, item["_event_d"]) if is_usd else Decimal("1.0")
        )
        dist = Distribution(
            id=str(uuid.uuid4()),
            workspace_id=job.workspace_id,
            financial_institution_id=fi_id,
            asset_id=item.get("asset_id"),
            type=item["_dist_type"],
            event_date=item["_event_d"],
            gross_amount=item["_gross"],
            tax=item["_tax"],
            net_amount=item["_net"],
            currency=item["_currency_enum"],
            fx_rate=dist_fx,
            notes=f"avenue proventos (job {job.id})",
            is_active=True,
            external_id=item["external_id"],
            external_source=ExternalSource.MANUAL_CSV,
            created_at=now, updated_at=now,
            created_by=user_id, updated_by=user_id,
        )
        db.add(dist)
        db.flush()
        created.append({
            "distribution_id": dist.id,
            "external_id": item["external_id"],
            "event_date": item["event_date"],
            "ticker": item["ticker"],
            "asset_id": item.get("asset_id"),
            "asset_name": item.get("asset_name"),
            "type": item["type"],
            "gross_amount": item["gross_amount"],
            "tax_amount": item["tax_amount"],
            "net_amount": item["net_amount"],
            "currency": item["currency"],
            "institution_short_name": item["institution_short_name"],
        })

    # Option premium events → AssetMovement SELL_OPEN / BUY_TO_CLOSE.
    for item in opt_plan:
        asset: Asset = item["_asset"]
        # Convenção AssetMovement (multicurrency_fx_rate_design): fx_rate
        # armazena PTAX do dia SEMPRE (BRL e USD). Consumer normaliza pra
        # multiplicador na hora de exibir (services/proventos.py:273,
        # _build_synthetic_premiums em distributions.py).
        mov_fx = resolve_fx_rate(db, item["_event_d"])
        mov = AssetMovement(
            id=str(uuid.uuid4()),
            workspace_id=job.workspace_id,
            asset_id=asset.id,
            type=item["_side_enum"],
            event_date=item["_event_d"],
            quantity=item["_qty"],
            unit_price=item["_unit_price"],
            gross_amount=item["_gross"],
            fee=item["_fee"],
            net_amount=item["_net"],
            currency=item["_currency_enum"],
            fx_rate=mov_fx,
            notes=f"{job.institution_id and 'XP' or ''} proventos opção (job {job.id})",
            is_active=True,
            external_id=item["external_id"],
            external_source=ExternalSource.MANUAL_CSV,
            created_at=now, updated_at=now,
            created_by=user_id, updated_by=user_id,
        )
        db.add(mov)
        db.flush()
        created.append({
            "movement_id": mov.id,
            "external_id": item["external_id"],
            "event_date": item["event_date"],
            "ticker": item["option_ticker"],
            "asset_id": asset.id,
            "asset_name": asset.name,
            "type": f"OPTION_{item['side']}",
            "gross_amount": item["gross_amount"],
            "tax_amount": None,
            "net_amount": item["net_amount"],
            "currency": item["currency"],
            "institution_short_name": item["institution_short_name"],
            "row_kind": "option_event",
        })

    detail = BulkApplyDetail(
        applied=created,
        matched_no_pendency=duplicate_rows + opt_duplicate,  # repurposed: duplicates
        orphan=orphan_rows + opt_orphan,
        pendency_not_in_extract=[],          # n/a for income
        auto_skipped=[],                     # n/a
    )
    return ExtractionApplyResult(
        applied_count=len(created),
        skipped_count=(
            len(duplicate_rows) + len(orphan_rows)
            + len(opt_duplicate) + len(opt_orphan)
        ),
        errors=errors,
        bulk_detail=detail,
    )


def preview_bulk_income(
    db: Session, *, job_id: str,
) -> ExtractionApplyResult:
    """Read-only preview of a BROKER_INCOME job — same shape as confirm
    would return, but no writes."""
    job = db.get(ExtractionJob, job_id)
    if job is None:
        raise ExtractionError(f"Job {job_id} not found")
    payload = job.extracted_json or {}
    apply_plan, preview_rows, orphan_rows, duplicate_rows, errors = (
        _classify_bulk_income(db, job, payload)
    )
    (
        opt_plan, opt_preview, opt_orphan, opt_duplicate, opt_errors,
    ) = _classify_bulk_option_events(db, job, payload)
    detail = BulkApplyDetail(
        applied=preview_rows + opt_preview,
        matched_no_pendency=duplicate_rows + opt_duplicate,
        orphan=orphan_rows + opt_orphan,
        pendency_not_in_extract=[],
        auto_skipped=[],
    )
    return ExtractionApplyResult(
        applied_count=len(preview_rows) + len(opt_preview),
        skipped_count=(
            len(orphan_rows) + len(duplicate_rows)
            + len(opt_orphan) + len(opt_duplicate)
        ),
        errors=errors + opt_errors,
        bulk_detail=detail,
    )
